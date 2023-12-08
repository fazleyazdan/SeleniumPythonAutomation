#  In this task we are going to write different data into excel file
#* if you want to do this we have to specify it 

import openpyxl

file_path = "D:\\for_writing_diff_data.xlsx"            #! save file path 
workbook = openpyxl.load_workbook(file_path)       #! load workbook
sheet = workbook.active                         
# sheet = workbook["data"]                         #! This can also be used                         


#* writing data into rows and col

sheet.cell(1,1).value = "1"
sheet.cell(1,2).value = "Ali"
sheet.cell(1,3).value = "Developer"


sheet.cell(2,1).value = "2"
sheet.cell(2,2).value = "Jawad"
sheet.cell(2,3).value = "QA"


sheet.cell(3,1).value = "3"
sheet.cell(3,2).value = "Kamran"
sheet.cell(3,3).value = "DevOps"

workbook.save(file_path)

