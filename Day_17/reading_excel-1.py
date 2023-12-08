
#* so we have an Excel file then inside we have workbooks then we have sheets inside it are multiple rows and inside rows we have multiple cells
#! File --> Workbooks --> Sheets --> Rows --> Cells
 
import openpyxl

file_path = "D:\for_reading_data.xlsx"             #! save file path 
workbook = openpyxl.load_workbook(file_path)       #! load workbook
sheet = workbook["Books"]                          #! save sheet 
# sheet = workbook.active                           #! use this if there is only one excel sheet

#* now we can print number of rows and columns 

rows = sheet.max_row                      
col = sheet.max_column

print("number of rows:", rows)
print("number of col:", col)

#* reading all rows and col data from excel

for r in range(1, rows+1):                        #! for rows (in Range F, last value is not counted so we had to add 1)
    for c in range(1, col+1):                     #! for col  
        print(sheet.cell(r, c).value, end="   ")  #! to extract data from cell we have to use sheet as it contains cells
    print()