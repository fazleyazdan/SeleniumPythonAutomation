import openpyxl
from openpyxl.styles import PatternFill

def getRowCount(file_path, sheet_name):
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook[sheet_name]
    return sheet.max_row

def getColumnCount(file_path, sheet_name):
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook[sheet_name]
    return sheet.max_column

def readData (file_path, sheet_name, rowno , colno):
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook[sheet_name]
    data = sheet.cell(rowno, colno).value
    return data

def writeData (file_path, sheet_name, rowno , colno, data):
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook[sheet_name]
    sheet.cell(rowno, colno).value =  data
    workbook.save(file_path)
    
def fillGreenColor(file_path, sheet_name, rowno, colno):        #! for green color
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook[sheet_name]
    greenFill = PatternFill(start_color='60b212' , end_color= '60b212', fill_type='solid')
    sheet.cell(rowno,colno).fill = greenFill
    workbook.save(file_path)
    
def fillRedColor(file_path, sheet_name, rowno, colno):        #! for red color
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook[sheet_name]
    redFill = PatternFill(start_color='ff0000' , end_color= 'ff0000', fill_type='solid')
    sheet.cell(rowno,colno).fill = redFill
    workbook.save(file_path)
    
    